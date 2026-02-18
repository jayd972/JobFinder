"""Spreadsheet import and export utilities."""

import os
import re
import logging
from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

# ── Column name mapping (flexible header accept) ───────────────────

COLUMN_MAP = {
    # Required
    "company_name": ["company_name", "company name", "company", "name"],
    "workday_url": ["workday_careers_url", "workday_url", "careers_url", "careers site",
                     "careers_site", "url", "workday_link", "link"],
    # Optional
    "note": ["note", "notes", "comment", "comments"],
    "preferred_keywords": ["preferred_keywords", "keywords", "preferred keywords"],
}


def _match_column(header):
    """Map a header string to a canonical column name."""
    h = header.strip().lower()
    for canonical, variants in COLUMN_MAP.items():
        if h in variants:
            return canonical
    return None


def import_companies(file_path):
    """
    Import companies from an xlsx file.

    Returns: list of dicts with company_name, workday_url, note, preferred_keywords
    """
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    # Read headers from first row
    headers = []
    for cell in ws[1]:
        val = str(cell.value or "").strip()
        canonical = _match_column(val)
        headers.append(canonical)

    if "company_name" not in headers:
        raise ValueError("Spreadsheet must have a 'company_name' or 'Company name' column")
    if "workday_url" not in headers:
        raise ValueError("Spreadsheet must have a 'workday_careers_url' or 'Careers site' column")

    companies = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue

        record = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                record[headers[i]] = str(val).strip() if val else None

        if record.get("company_name") and record.get("workday_url"):
            # Generate a stable company_id from the name
            cid = re.sub(r"[^a-z0-9]", "_", record["company_name"].lower())[:20]
            record["company_id"] = cid
            companies.append(record)

    wb.close()
    return companies


def export_results(jobs, file_path):
    """
    Export job results to an xlsx file.

    Columns: Company, Title, Location, Posted, Job Link, Match Score, Why Matched
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Results"

    # Headers
    headers = ["Company", "Title", "Location", "Posted", "Job Link", "Match Score", "Why Matched"]
    ws.append(headers)

    # Style headers
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for job in jobs:
        ws.append([
            job.get("company_name", ""),
            job.get("title", ""),
            job.get("locations_text", ""),
            job.get("posted_label", ""),
            job.get("job_url", ""),
            job.get("match_score", 0),
            job.get("matched_keywords", ""),
        ])

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(file_path)
    logger.info(f"Exported {len(jobs)} jobs to {file_path}")
    return file_path
